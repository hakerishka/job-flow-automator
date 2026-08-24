import sys
import os
import re
import glob
import time
import subprocess
from datetime import datetime
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import streamlit as st
from google import genai
from google.genai import types
from google.genai.errors import APIError

import config

# --- PAGE CONFIGURATION ---
st.set_page_config(
    page_title="Job-Flow Automator",
    page_icon="🎯",
    layout="wide"
)

BASE_DIR = Path(__file__).parent.resolve()
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

# Path mapping for EN and DE assets
MASTER_CV_EN = BASE_DIR / "Master_CV.md"
MASTER_CV_DE = BASE_DIR / "Master_CV_DE.md"
MAIN_TYP_EN = BASE_DIR / "main.typ"
MAIN_TYP_DE = BASE_DIR / "main_de.typ"
TAILORED_TYP_PATH = BASE_DIR / "tailored.typ"
REVIEWED_XLSX_PATH = BASE_DIR / "reviewed.xlsx"

# Fallback to templates if local active files don't exist yet
if not MASTER_CV_EN.exists() and (BASE_DIR / "templates" / "master_cv.template.md").exists():
    MASTER_CV_EN = BASE_DIR / "templates" / "master_cv.template.md"
if not MASTER_CV_DE.exists() and (BASE_DIR / "templates" / "master_cv_de.template.md").exists():
    MASTER_CV_DE = BASE_DIR / "templates" / "master_cv_de.template.md"

if not MAIN_TYP_EN.exists() and (BASE_DIR / "templates" / "main.template.typ").exists():
    MAIN_TYP_EN = BASE_DIR / "templates" / "main.template.typ"
if not MAIN_TYP_DE.exists() and (BASE_DIR / "templates" / "main_de.template.typ").exists():
    MAIN_TYP_DE = BASE_DIR / "templates" / "main_de.template.typ"


# --- HELPER FUNCTIONS FOR REVIEW TRACKING ---
def mark_job_in_reviewed_file(job_url, title, company, status="Rejected", color="YELLOW"):
    """Append or update a job in reviewed.xlsx with background color and status."""
    clean_url = str(job_url).split('?')[0].rstrip('/') if job_url else ""
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    fills = {
        "YELLOW": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "BLUE": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
        "GREEN": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    }
    fill = fills.get(color, fills["YELLOW"])

    if not REVIEWED_XLSX_PATH.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reviewed Jobs"
        ws.append(["status", "title", "company", "job_url", "timestamp"])
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
    else:
        wb = openpyxl.load_workbook(REVIEWED_XLSX_PATH)
        ws = wb.active

    existing_urls = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= 4 and row[3]:
            existing_urls.add(str(row[3]).split('?')[0].rstrip('/'))

    if clean_url and clean_url not in existing_urls:
        new_row_idx = ws.max_row + 1
        ws.append([status, title, company, clean_url, now_str])
        for col_idx in range(1, 6):
            ws.cell(row=new_row_idx, column=col_idx).fill = fill
        wb.save(REVIEWED_XLSX_PATH)
        return True
    return False


def get_all_reviewed_urls():
    """Retrieve set of all marked URLs across all .xlsx files in project and history/."""
    marked_urls = set()
    xlsx_files = glob.glob(str(BASE_DIR / "*.xlsx")) + glob.glob(str(BASE_DIR / "history" / "*.xlsx"))
    for xlsx_file in set(xlsx_files):
        try:
            wb = openpyxl.load_workbook(xlsx_file, data_only=True)
            ws = wb.active
            url_col_idx = None
            for col_idx, cell in enumerate(ws[1], 1):
                if str(cell.value).strip().lower() == 'job_url':
                    url_col_idx = col_idx
                    break
            if url_col_idx:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and len(row) >= url_col_idx and row[url_col_idx - 1]:
                        clean_u = str(row[url_col_idx - 1]).split('?')[0].rstrip('/')
                        marked_urls.add(clean_u)
        except Exception:
            pass
    return marked_urls


# --- SIDEBAR: SETTINGS & MODEL SELECTION ---
st.sidebar.title("🎯 Job-Flow Control")

api_key_env = os.environ.get("GEMINI_API_KEY", "")
api_key_input = st.sidebar.text_input(
    "Gemini API Key:",
    value=api_key_env,
    type="password",
    help="Ключ считывается из переменной GEMINI_API_KEY или вводится вручную."
)

selected_model = st.sidebar.selectbox(
    "Модель Gemini:",
    options=[
        "gemini-3.7-flash",
        "gemini-3.5-flash-lite",
        "gemini-3.5-flash"
    ],
    index=0
)

# Status indicators
st.sidebar.divider()
st.sidebar.caption("📁 Статус локальных файлов:")
st.sidebar.write(f"• Master CV (EN): {'✅ Найден' if MASTER_CV_EN.exists() else '❌ Не найден'}")
st.sidebar.write(f"• Master CV (DE): {'✅ Найден' if MASTER_CV_DE.exists() else '❌ Не найден'}")
st.sidebar.write(f"• Typst Main (EN): {'✅ Найден' if MAIN_TYP_EN.exists() else '❌ Не найден'}")
st.sidebar.write(f"• Typst Main (DE): {'✅ Найден' if MAIN_TYP_DE.exists() else '❌ Не найден'}")

reviewed_count = len(get_all_reviewed_urls())
st.sidebar.write(f"• Отслежено вакансий: **{reviewed_count}** шт.")


# --- SESSION STATE ---
if "chat" not in st.session_state:
    st.session_state.chat = None
if "phase_1_response" not in st.session_state:
    st.session_state.phase_1_response = None
if "jd_text" not in st.session_state:
    st.session_state.jd_text = ""
if "target_role" not in st.session_state:
    st.session_state.target_role = ""
if "target_lang" not in st.session_state:
    st.session_state.target_lang = "English"


# --- GEMINI CLIENT ---
@st.cache_resource(show_spinner=False)
def get_client(api_key: str):
    return genai.Client(api_key=api_key) if api_key else None

client = get_client(api_key_input)


def get_system_instruction(target_lang="English"):
    """Dynamically generate system prompt based on selected resume language."""
    cv_path = MASTER_CV_DE if target_lang == "Deutsch" and MASTER_CV_DE.exists() else MASTER_CV_EN
    cv_content = ""
    if cv_path.exists():
        with open(cv_path, "r", encoding="utf-8") as f:
            cv_content = f.read()

    lang_rule = (
        "Execute Phase 1 strictly in RUSSIAN. Execute Phase 2 strictly in GERMAN (valid Typst code, high-density professional German ATS terminology, e.g. '01/2023 – Heute', 'Freiberuflich / Selbstständig')."
        if target_lang == "Deutsch"
        else "Execute Phase 1 strictly in RUSSIAN. Execute Phase 2 strictly in ENGLISH (valid Typst code)."
    )

    return f"""
<role>
You are an expert ATS Optimization Architect and Lead Technical Recruiter specializing in European tech hubs (Berlin, Munich, Zurich). Your task is to analyze a candidate's Master CV against a target Job Description (JD) through a rigorous two-phase interactive process and generate precise Typst variable declarations.
</role>

<rules>
1. LANGUAGE RULE: {lang_rule}
2. Strictly execute Phase 1 first. Do NOT generate Phase 2 until the user responds to Phase 1 clarification questions.
3. Maintain a strict 1-page total document budget. Select only the 2-3 most impactful bullet points per role.
4. ATS DELIMITERS: NEVER use vertical pipes ("|"). Use only standard middle dots (" · ") or commas for separators.
5. ENCODING: Use only standard ASCII hyphens ("-") in text. Never use unicode non-breaking hyphens.
6. Output in Phase 2 MUST be strictly valid Typst code containing ONLY the variable declarations (`#let target-role = ...`, `#let summary = [...]`, `#let skills = [...]`, `#let experience = [...]`).
</rules>

<master_cv>
{cv_content}
</master_cv>
"""


def send_message_with_retry(chat_session, prompt_text, max_retries=3):
    for attempt in range(max_retries):
        try:
            return chat_session.send_message(prompt_text)
        except APIError as e:
            if "503" in str(e) or "UNAVAILABLE" in str(e):
                if attempt < max_retries - 1:
                    sleep_time = 2 * (attempt + 1)
                    st.info(f"Сервер временно занят. Повторный запрос через {sleep_time} сек...")
                    time.sleep(sleep_time)
                    continue
            raise e
        except Exception as e:
            raise e


# --- MAIN TABS ---
tab_feed, tab_tailor, tab_scraper, tab_guide = st.tabs([
    "📋 1. Лента вакансий & Трекер",
    "⚡ 2. ATS Resume Tailor",
    "🚀 3. Запуск сбора вакансий",
    "📖 4. Инструкция & Настройки"
])


# ==============================================================================
# TAB 1: JOB FEED & REVIEW TRACKER
# ==============================================================================
with tab_feed:
    st.subheader("📋 Лента собранных вакансий")
    
    live_stream_file = BASE_DIR / "jobs_live_stream.csv"
    if live_stream_file.exists():
        mtime = os.path.getmtime(live_stream_file)
        if time.time() - mtime < 1800:
            c_live1, c_live2 = st.columns([4, 1])
            with c_live1:
                st.info("🟢 **Live-поток активен:** вакансии поступают в реальном времени! Вы можете сразу откликаться или скрывать неподходящие.")
            with c_live2:
                if st.button("🔄 Обновить список", use_container_width=True):
                    st.rerun()

    csv_files = []
    if live_stream_file.exists():
        csv_files.append(str(live_stream_file))
    
    scraped_csvs = glob.glob(str(BASE_DIR / "jobs_clean_*.csv")) + glob.glob(str(BASE_DIR / "berlin_jobs_clean_*.csv")) + glob.glob(str(BASE_DIR / "history" / "*.csv"))
    scraped_csvs = sorted(list(set(scraped_csvs)), key=os.path.getctime, reverse=True)
    
    for f in scraped_csvs:
        if f not in csv_files:
            csv_files.append(f)

    if not csv_files:
        st.info("ℹ️ В папке пока нет файлов выгрузки (`jobs_clean_*.csv`). Перейдите на вкладку **«🚀 Запуск сбора вакансий»** для первого запуска.")
    else:
        def format_csv_name(path_str):
            name = os.path.basename(path_str)
            if name == "jobs_live_stream.csv":
                return "🔴 [LIVE STREAM] Текущий живой поток"
            return name

        col_f1, col_f2, col_f3 = st.columns([3, 2, 2])
        with col_f1:
            selected_csv = st.selectbox("Выберите файл выгрузки:", csv_files, format_func=format_csv_name)
        with col_f2:
            hide_reviewed = st.checkbox("Скрыть уже отработанные (Reviewed)", value=True)
        with col_f3:
            min_score = st.slider("Мин. Match Score:", 0, 100, 20, step=5)

        try:
            df = pd.read_csv(selected_csv)
            reviewed_urls = get_all_reviewed_urls()

            df['clean_url'] = df['job_url'].apply(lambda u: str(u).split('?')[0].rstrip('/') if pd.notna(u) else "")
            df['is_reviewed'] = df['clean_url'].isin(reviewed_urls)

            # Filtering
            filtered_df = df.copy()
            if hide_reviewed:
                filtered_df = filtered_df[~filtered_df['is_reviewed']]
            if 'match_score' in filtered_df.columns:
                filtered_df = filtered_df[filtered_df['match_score'] >= min_score]

            # Category filter
            categories = ["Все категории"] + sorted(list(filtered_df['category'].dropna().unique())) if 'category' in filtered_df.columns else []
            selected_cat = st.selectbox("Фильтр по категории:", categories)
            if selected_cat != "Все категории":
                filtered_df = filtered_df[filtered_df['category'] == selected_cat]

            st.caption(f"Найдено: **{len(filtered_df)}** вакансий (из {len(df)} исходных)")

            # Render Cards
            for idx, row in filtered_df.iterrows():
                title = row.get('title', 'Unknown Title')
                company = row.get('company', 'Unknown Company')
                category = row.get('category', 'General')
                score = row.get('match_score', 0)
                url = row.get('job_url', '')
                date_posted = row.get('date_posted', '')
                desc = str(row.get('description', ''))
                email = row.get('contact_email', '')

                with st.container(border=True):
                    c_title, c_score = st.columns([5, 1])
                    with c_title:
                        st.markdown(f"### {title}")
                        st.markdown(f"**🏢 {company}** · 📂 `{category}` · 📅 {date_posted}")
                        if email and pd.notna(email):
                            st.caption(f"📧 Контакт: `{email}`")
                    with c_score:
                        st.metric("Match", f"{score}%")

                    # Action buttons in one clean line
                    b_col1, b_col2, b_col3, b_col4, b_col5 = st.columns([2, 2, 2, 2, 2])
                    
                    with b_col1:
                        if url:
                            st.link_button("🔗 Открыть", url, help="Открыть страницу вакансии в новой вкладке")
                        else:
                            st.button("Нет ссылки", disabled=True, key=f"nourl_{idx}")

                    with b_col2:
                        if st.button("🇬🇧 CV (EN)", key=f"tailor_en_{idx}", type="primary", help="Адаптировать резюме на английском"):
                            st.session_state.jd_text = desc
                            st.session_state.target_role = title
                            st.session_state.target_lang = "English"
                            st.session_state.phase_1_response = None
                            st.session_state.chat = None
                            st.toast(f"Вакансия '{title}' (EN) скопирована в ATS Tailor!", icon="🇬🇧")

                    with b_col3:
                        if st.button("🇩🇪 CV (DE)", key=f"tailor_de_{idx}", help="Адаптировать резюме на немецком"):
                            st.session_state.jd_text = desc
                            st.session_state.target_role = title
                            st.session_state.target_lang = "Deutsch"
                            st.session_state.phase_1_response = None
                            st.session_state.chat = None
                            st.toast(f"Вакансия '{title}' (DE) скопирована в ATS Tailor!", icon="🇩🇪")

                    with b_col4:
                        if st.button("🟡 Пропуск", key=f"reject_{idx}", help="Пометить желтым и скрыть"):
                            mark_job_in_reviewed_file(url, title, company, status="Rejected", color="YELLOW")
                            st.toast(f"Отмечено как 'Не подходит': {title}", icon="🟡")
                            st.rerun()

                    with b_col5:
                        if st.button("🔵 Отклик", key=f"apply_{idx}", help="Пометить синим как 'Откликнулся'"):
                            mark_job_in_reviewed_file(url, title, company, status="Applied", color="BLUE")
                            st.toast(f"Отмечено как 'Отклик': {title}", icon="🔵")
                            st.rerun()

                    with st.expander("👁️ Посмотреть описание вакансии"):
                        st.write(desc)

        except Exception as e:
            st.error(f"Ошибка загрузки CSV: {e}")


# ==============================================================================
# TAB 2: ATS RESUME TAILOR ENGINE
# ==============================================================================
with tab_tailor:
    st.subheader("⚡ Интерактивная адаптация резюме (ATS Engine)")

    c_lang, c_space = st.columns([2, 3])
    with c_lang:
        selected_lang = st.radio(
            "Язык итогового резюме:",
            options=["English", "Deutsch"],
            index=0 if st.session_state.target_lang == "English" else 1,
            horizontal=True
        )
        st.session_state.target_lang = selected_lang

    active_cv_path = MASTER_CV_DE if selected_lang == "Deutsch" and MASTER_CV_DE.exists() else MASTER_CV_EN
    active_typ_path = MAIN_TYP_DE if selected_lang == "Deutsch" and MAIN_TYP_DE.exists() else MAIN_TYP_EN

    if not api_key_input:
        st.warning("⚠️ Для генерации укажите Gemini API Key в боковой панели слева.")
    elif not active_cv_path.exists():
        st.error(f"❌ Не найден файл `{active_cv_path.name}`.")
    elif not active_typ_path.exists():
        st.error(f"❌ Не найден файл `{active_typ_path.name}`.")
    else:
        jd_input = st.text_area(
            "Текст вакансии (Job Description):",
            value=st.session_state.jd_text,
            height=180,
            placeholder="Вставьте описание вакансии или нажмите кнопку адаптации в Ленте вакансий..."
        )

        col1, col2 = st.columns([1, 5])
        with col1:
            analyze_btn = st.button("🔍 Анализ вакансии (Фаза 1)", type="primary", use_container_width=True)
        with col2:
            if st.button("🔄 Сбросить"):
                st.session_state.chat = None
                st.session_state.phase_1_response = None
                st.session_state.jd_text = ""
                st.rerun()

        if analyze_btn and jd_input.strip():
            st.session_state.jd_text = jd_input
            with st.spinner("Анализирую соответствие Master CV требованиям вакансии..."):
                try:
                    sys_inst = get_system_instruction(selected_lang)
                    st.session_state.chat = client.chats.create(
                        model=selected_model,
                        config=types.GenerateContentConfig(
                            system_instruction=sys_inst,
                            temperature=0.2,
                        )
                    )

                    phase_1_prompt = f"""
                    <job_description>
                    {jd_input}
                    </job_description>

                    --- PHASE 1: GAP ANALYSIS & STRATEGIC CLARIFICATION (НА РУССКОМ ЯЗЫКЕ) ---

                    Сравни <master_cv> и <job_description>. Выведи СТРОГО следующие два блока на русском языке:

                    1. **ATS Анализ соответствия ({'на немецком языке' if selected_lang == 'Deutsch' else 'на английском языке'}):**
                       - Оценка совпадения (в % от 0 до 100%).
                       - Совпавшие ключевые слова (навыки, инструменты и процессы из вакансии, которые уже есть в CV).
                       - Критические пробелы, если есть (требования вакансии, которые отсутствуют или слабо выражены).

                    2. **Уточняющие вопросы (максимум 3 вопроса):**
                       - Вопрос по недостающему софту/инструментам (уточнить, пишем ли "Working knowledge of [Инструмент]" / "Gute Kenntnisse in [Tool]").
                       - Вопрос по адаптации софт-скиллов и формулировок под специфику роли.
                       - Вопрос по тональности (строго корпоративная [Corporate-Safe] или стартап-профиль).

                    Заверши Фазу 1 точной фразой:
                    "Ответьте на эти вопросы, чтобы я сгенерировал код переменных для tailored.typ."
                    """
                    response = send_message_with_retry(st.session_state.chat, phase_1_prompt)
                    st.session_state.phase_1_response = response.text
                except Exception as e:
                    st.error(f"Ошибка Gemini API: {e}")

        # PHASE 2: Clarification & PDF Compilation
        if st.session_state.phase_1_response:
            st.divider()
            st.markdown(st.session_state.phase_1_response)

            user_answers = st.text_area(
                "Ваши ответы на вопросы (кратко):",
                height=100,
                placeholder="1. Да, добавь working knowledge of X. 2. Опыт трансляций подать как hardware uptime. 3. Профиль стартап."
            )

            if st.button("🚀 Сгенерировать и скомпилировать 1-Page PDF", type="primary"):
                with st.spinner("Генерация Typst переменных и компиляция PDF..."):
                    try:
                        summary_placeholder = (
                            "ZUSAMMENFASSUNG_PROFIL (maximal 3-4 Zeilen, hohe Dichte relevanter deutscher ATS-Schlüsselwörter, nur ASCII-Bindestriche)."
                            if selected_lang == "Deutsch"
                            else "SUMMARY_PARAGRAPH (3-4 lines maximum, high keyword density, standard ASCII hyphens only)."
                        )

                        skills_header = (
                            "- *Kernkompetenzen & Systeme:* Keyword 1, Keyword 2, Keyword 3, Keyword 4\n  - *Tools & Plattformen:* Tool 1, Tool 2, Tool 3, Tool 4\n  - *Methoden & Standards:* Method 1, Method 2, Method 3"
                            if selected_lang == "Deutsch"
                            else "- *Core Technical & Systems:* Keyword 1, Keyword 2, Keyword 3, Keyword 4\n  - *Tools & Platforms:* Tool 1, Tool 2, Tool 3, Tool 4\n  - *Operational Methodologies:* Method 1, Method 2, Method 3"
                        )

                        exp_example = (
                            "*Positionsbezeichnung 1* — _Unternehmen 1_ #h(1fr) #text(fill: rgb(\"#444444\"), size: 8.5pt)[01/2023 – Heute · Berlin / Remote]\n  - Stichpunkt 1 (Aktionsverb + Kontext + Messbares Ergebnis)\n  - Stichpunkt 2\n  - Stichpunkt 3\n\n  #v(0.25em)\n  *Positionsbezeichnung 2* — _Unternehmen 2_ #h(1fr) #text(fill: rgb(\"#444444\"), size: 8.5pt)[09/2017 – 01/2023 · Standort]\n  - Stichpunkt 1\n  - Stichpunkt 2\n  - Stichpunkt 3"
                            if selected_lang == "Deutsch"
                            else "*Role Title 1* — _Company 1_ #h(1fr) #text(fill: rgb(\"#444444\"), size: 8.5pt)[Jan 2023 – Present · Berlin / Remote]\n  - Bullet point 1\n  - Bullet point 2\n  - Bullet point 3\n\n  #v(0.25em)\n  *Role Title 2* — _Company 2_ #h(1fr) #text(fill: rgb(\"#444444\"), size: 8.5pt)[Sep 2017 – Jan 2023 · Location]\n  - Bullet point 1\n  - Bullet point 2\n  - Bullet point 3"
                        )

                        phase_2_prompt = f"""
                        User Answers:
                        {user_answers if user_answers.strip() else "Proceed with standard optimal mappings based on Master CV."}

                        --- PHASE 2: TYPST VARIABLE GENERATION ({'IN GERMAN' if selected_lang == 'Deutsch' else 'IN ENGLISH'}) ---

                        Generate the exact Typst variable block ready to paste directly into `tailored.typ`:

                        #let target-role = "{st.session_state.target_role if st.session_state.target_role else 'EXACT_JOB_TITLE_FROM_JD'}"

                        #let summary = [
                          {summary_placeholder}
                        ]

                        #let skills = [
                          {skills_header}
                        ]

                        #let experience = [
                          {exp_example}
                        ]
                        """
                        response = send_message_with_retry(st.session_state.chat, phase_2_prompt)
                        raw_code = response.text

                        # Clean code block delimiters
                        cleaned_code = re.sub(r"^```typst\s*|^```\s*|```$", "", raw_code.strip(), flags=re.MULTILINE)
                        cleaned_code = cleaned_code.replace("\u2011", "-").replace("\u2013", "-").replace("\u2014", "—")
                        cleaned_code = cleaned_code.replace(" | ", " · ")

                        # Write to tailored.typ
                        with open(TAILORED_TYP_PATH, "w", encoding="utf-8") as f:
                            f.write(cleaned_code)

                        # Output file naming
                        role_match = re.search(r'#let target-role = "(.*?)"', cleaned_code)
                        role_title = role_match.group(1) if role_match else "Custom_Role"
                        role_slug = re.sub(r'[\\/*?:"<>|]', "", role_title).replace(" ", "_").replace("/", "-")

                        date_str = datetime.now().strftime("%Y-%m-%d")
                        lang_suffix = "DE" if selected_lang == "Deutsch" else "EN"
                        output_pdf_name = f"CV_Tailored_{role_slug}_{date_str}_{lang_suffix}.pdf"
                        output_pdf_path = OUTPUT_DIR / output_pdf_name

                        # Compile Typst using active template
                        compile_res = subprocess.run(
                            ["typst", "compile", str(active_typ_path), str(output_pdf_path)],
                            capture_output=True,
                            text=True
                        )

                        if compile_res.returncode == 0:
                            st.success(f"✅ Резюме успешно скомпилировано в 1 страницу ({selected_lang}): `{output_pdf_name}`")
                            with open(output_pdf_path, "rb") as pdf_file:
                                st.download_button(
                                    label=f"📥 Скачать готовый PDF ({selected_lang})",
                                    data=pdf_file.read(),
                                    file_name=output_pdf_name,
                                    mime="application/pdf"
                                )
                            with st.expander("Посмотреть сгенерированный код `tailored.typ`"):
                                st.code(cleaned_code, language="typst")
                        else:
                            st.error("Ошибка Typst CLI при компиляции:")
                            st.code(compile_res.stderr)
                            st.info("💡 Убедитесь, что утилита `typst` установлена в вашей системе (`winget install --id Typst.Typst` или скачайте с typst.app).")

                    except Exception as e:
                        st.error(f"Ошибка API при выполнении Фазы 2: {e}")


# ==============================================================================
# TAB 3: JOB SCRAPER
# ==============================================================================
with tab_scraper:
    st.subheader("🚀 Запуск сбора вакансий")
    st.write("Сбор данных из LinkedIn, Indeed, VC Портфолио (Ashby, Greenhouse), Arbeitnow и Berlin Startup Jobs.")

    s_col1, s_col2 = st.columns(2)
    with s_col1:
        st.write(f"• Локация: **{config.LOCATION}**")
        st.write(f"• Глубина поиска: **{config.HOURS_OLD} ч. ({config.HOURS_OLD // 24} дн.)**")
        st.write(f"• Вакансий на запрос: **{config.RESULTS_PER_QUERY}**")
    with s_col2:
        st.write(f"• Категорий поиска: **{len(config.SEARCH_CATEGORIES)}**")
        st.write(f"• Исключение языка: **{len(config.ACTIVE_LANGUAGE_PATTERNS)} regex-правил**")

    if st.button("▶️ Запустить полный цикл сбора", type="primary"):
        st.write("🚀 **Запуск скрейпера в реальном времени...**")
        log_box = st.empty()
        logs = []
        
        sub_env = {
            **os.environ,
            "PYTHONIOENCODING": "utf-8",
            "PYTHONUTF8": "1",
            "PYTHONUNBUFFERED": "1"
        }
        
        proc = subprocess.Popen(
            [sys.executable, "-u", str(BASE_DIR / "main.py")],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
            env=sub_env
        )
        
        for line in iter(proc.stdout.readline, ''):
            if line:
                logs.append(line)
                # Keep last 22 lines streaming in browser
                log_box.code("".join(logs[-22:]), language="text")
                
        proc.stdout.close()
        returncode = proc.wait()
        
        if returncode == 0:
            st.success("🎉 Сбор успешно завершен! Свежий файл сохранен. Перейдите на вкладку '📋 1. Лента вакансий' для просмотра.")
        else:
            st.error("❌ Сбой при сборе вакансий. Ознакомьтесь с логом ниже.")

        with st.expander("📜 Посмотреть полный лог выполнения (можно скопировать весь текст целиком)", expanded=False):
            st.text_area("Полный лог терминала:", value="".join(logs), height=350)


# ==============================================================================
# TAB 4: GUIDE & SETTINGS
# ==============================================================================
with tab_guide:
    st.subheader("📖 Руководство пользователя и Настройки")
    
    st.markdown("""
    ### 🔄 Рабочий процесс полного цикла (Workflow):
    1. **Сбор вакансий:** Запустите `main.py` (или нажмите кнопку во вкладке 3).
    2. **Отбор и проверка:** Во вкладке **«Лента вакансий»** открывайте ссылки в 1 клик. Если вакансия не подходит — жмите 🟡 `Пропуск` (она больше не появится).
    3. **Генерация резюме (EN или DE):**
       - Нажмите 🇬🇧 `CV (EN)` или 🇩🇪 `CV (DE)` прямо в карточке вакансии.
       - Ответьте на 3 уточняющих вопроса в Фазе 1.
       - Скачайте идеальный 1-страничный PDF на нужном языке.
    
    ---
    ### 📂 Структура файлов резюме:
    - **`Master_CV.md` / `Master_CV_DE.md`** — Банк опыта (на английском и немецком языках).
    - **`main.typ` / `main_de.typ`** — Скелеты резюме Typst с заголовками и статичными блоками.
    - **`tailored.typ`** — Динамический файл, куда AI записывает адаптированные под вакансию summary, skills и bullet points.
    - **`templates/`** — Обезличенные шаблоны для новых пользователей на GitHub.
    """)
