"""
🚀 MAIN JOB AGGREGATOR & SMART FILTER PIPELINE (High-Performance & Live Streaming)
Automated workflow:
  1. Multi-source scraping (LinkedIn, Indeed + VC Portfolios via custom_boards)
  2. Instant on-the-fly deduplication against ALL historical reviewed files (< 1ms)
  3. Fast multithreaded LinkedIn description enrichment (5x speedup via ThreadPoolExecutor)
  4. Language requirement regex filtering (drops local language constraints)
  5. Email extraction & Profile Match scoring
  6. Live stream updating (jobs_live_stream.csv) + final timestamped CSV export
"""

import sys
import os
import re
import time
import random
import glob
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

# Ensure standard UTF-8 stream output on Windows
if hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

import pandas as pd
import openpyxl
import requests
from bs4 import BeautifulSoup
from jobspy import scrape_jobs

import config
from custom_boards import fetch_all_custom_boards

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LIVE_STREAM_PATH = os.path.join(BASE_DIR, "jobs_live_stream.csv")


def clean_job_url(url):
    """Strip query parameters and trailing slashes from URLs for reliable deduplication."""
    return str(url).split('?')[0].rstrip('/') if pd.notna(url) else ""


def normalize_text_for_dedup(text):
    """Normalize text by removing gender markers (m/f/d), legal entities (GmbH, Inc), and special characters."""
    if not isinstance(text, str):
        return ""
    t = text.lower()
    t = re.sub(r'\(m/f/d\)|\(m/w/d\)|\(f/m/d\)|\(all genders\)|\(gn\)', '', t)
    t = re.sub(r'\b(gmbh|inc|se|ag|ltd|corp|llc|co)\b', '', t)
    return re.sub(r'[^a-z0-9]', '', t)


def is_suitable_title(title):
    """Check if title does not contain any forbidden/unwanted keywords."""
    t = str(title).lower()
    return not any(bad in t for bad in config.EXCLUDE_TITLE_KEYWORDS)


def requires_excluded_language(text):
    """Check if job description requires local language proficiency based on regex patterns."""
    text_lower = str(text).lower()
    for pattern in config.ACTIVE_LANGUAGE_PATTERNS:
        if re.search(pattern, text_lower, re.DOTALL):
            return True
    return False


def extract_emails(text):
    """Extract contact emails from job descriptions, excluding image extensions."""
    if not isinstance(text, str):
        return ""
    emails = set(re.findall(config.EMAIL_REGEX, text))
    clean_emails = [e for e in emails if not e.lower().endswith(('.png', '.jpg', '.jpeg', '.svg', '.webp'))]
    return ", ".join(clean_emails)


def calculate_match_score(row):
    """Compute 0-100% profile relevance score using multi-track evaluation."""
    res = config.evaluate_job_match(row.get('title', ''), row.get('description', ''))
    return res['match_score']


def apply_job_scoring(df):
    """Calculate multi-track match scores, tracks, and skill tags for a DataFrame."""
    if df.empty:
        return df
    
    scores = []
    tracks = []
    skills = []
    
    for idx, row in df.iterrows():
        eval_res = config.evaluate_job_match(row.get('title', ''), row.get('description', ''))
        scores.append(eval_res['match_score'])
        tracks.append(eval_res['matched_track'])
        skills.append(eval_res['matched_skills'])
        
    df['match_score'] = scores
    df['matched_track'] = tracks
    df['matched_skills'] = skills
    return df


def load_all_historical_reviewed_keys():
    """Load all marked URLs and (Title+Company) keys from all local and historical spreadsheets."""
    marked_urls = set()
    marked_title_company = set()
    
    xlsx_files = glob.glob(os.path.join(BASE_DIR, "*.xlsx")) + glob.glob(os.path.join(BASE_DIR, "history", "*.xlsx"))
    
    for f in set(xlsx_files):
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
            url_col = None
            title_col = None
            comp_col = None
            for idx, cell in enumerate(ws[1], 1):
                h = str(cell.value).strip().lower() if cell.value else ""
                if h == 'job_url': url_col = idx
                elif h == 'title': title_col = idx
                elif h == 'company': comp_col = idx

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row:
                    u = clean_job_url(row[url_col - 1]) if url_col and len(row) >= url_col else ""
                    if u: marked_urls.add(u)
                    
                    t = str(row[title_col - 1]) if title_col and len(row) >= title_col and row[title_col - 1] else ""
                    c = str(row[comp_col - 1]) if comp_col and len(row) >= comp_col and row[comp_col - 1] else ""
                    if t and c:
                        norm_k = f"{normalize_text_for_dedup(t)}_{normalize_text_for_dedup(c)}"
                        if norm_k != "_": marked_title_company.add(norm_k)
        except Exception:
            pass
            
    return marked_urls, marked_title_company


USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0"
]


def fetch_single_linkedin_desc(url, max_retries=3):
    """Fetch full 'About the job' text for a LinkedIn listing using guest endpoint with backoff."""
    if not url or pd.isna(url):
        return ""
    
    job_id_match = re.search(r'/view/(\d+)', str(url)) or re.search(r'(\d{7,})', str(url))
    
    target_urls = []
    if job_id_match:
        jid = job_id_match.group(1)
        target_urls.append(f"https://www.linkedin.com/jobs-guest/jobs/api/jobPosting/{jid}")
    target_urls.append(str(url))

    for target_url in target_urls:
        for attempt in range(max_retries):
            try:
                headers = {
                    'User-Agent': random.choice(USER_AGENTS),
                    'Accept-Language': 'en-US,en;q=0.9,de;q=0.8',
                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                }
                res = requests.get(target_url, headers=headers, timeout=10)
                if res.status_code == 200 and res.text:
                    soup = BeautifulSoup(res.text, 'html.parser')
                    desc_div = (
                        soup.find('div', class_='show-more-less-html__markup') or 
                        soup.find('section', class_='description') or
                        soup.find('div', class_='description__text') or
                        soup.find(class_=lambda c: c and 'description' in str(c))
                    )
                    if desc_div:
                        desc_text = desc_div.get_text(separator='\n', strip=True)
                        if len(desc_text) > 80:
                            return desc_text
                elif res.status_code in [429, 503]:
                    backoff = random.uniform(1.5, 3.0) * (attempt + 1)
                    time.sleep(backoff)
                    continue
            except Exception:
                time.sleep(1.0)
    return ""


def enrich_descriptions_multithreaded(df):
    """Paced, polite enrichment of LinkedIn descriptions with retry passes and micro-delays."""
    df['description'] = df['description'].fillna('')
    df['site'] = df['site'].fillna('')
    
    mask = (df['site'].str.lower() == 'linkedin') & (df['description'].str.len() < 100)
    linkedin_indices = list(df[mask].index)
    
    if not linkedin_indices:
        return df

    print(f"\n🔄 [PACED-ENRICH] Concurrently fetching {len(linkedin_indices)} LinkedIn descriptions with smart rate-limiting...", flush=True)
    
    # Pass 1: Controlled concurrency (3 workers) with micro-delays
    def worker_fetch(idx):
        time.sleep(random.uniform(0.2, 0.5))
        return idx, fetch_single_linkedin_desc(df.at[idx, 'job_url'])

    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = [executor.submit(worker_fetch, i) for i in linkedin_indices]
        completed_count = 0
        success_count = 0
        for future in as_completed(futures):
            try:
                i, desc = future.result()
                if desc:
                    df.at[i, 'description'] = desc
                    success_count += 1
            except Exception:
                pass
            completed_count += 1
            if completed_count % 25 == 0 or completed_count == len(linkedin_indices):
                print(f"   • Pass 1: Progress {completed_count}/{len(linkedin_indices)} (Filled: {success_count})...", flush=True)

    # Pass 2: Retry remaining empty descriptions with spacing
    rem_mask = (df['site'].str.lower() == 'linkedin') & (df['description'].str.len() < 100)
    rem_indices = list(df[rem_mask].index)
    if rem_indices:
        print(f"   • Pass 2: Retrying {len(rem_indices)} uncollected descriptions with gentle spacing...", flush=True)
        time.sleep(2.0)
        for count, idx in enumerate(rem_indices, 1):
            time.sleep(random.uniform(0.5, 1.0))
            desc = fetch_single_linkedin_desc(df.at[idx, 'job_url'], max_retries=2)
            if desc:
                df.at[idx, 'description'] = desc
                success_count += 1
            if count % 15 == 0 or count == len(rem_indices):
                print(f"     -> Retried {count}/{len(rem_indices)}...", flush=True)

    print(f"  ✓ Description enrichment complete: {success_count}/{len(linkedin_indices)} filled.", flush=True)
    return df


def update_live_stream_file(df_accumulated):
    """Write current deduplicated and processed jobs to jobs_live_stream.csv."""
    if df_accumulated.empty:
        return
    existing_cols = [c for c in config.OUTPUT_COLUMNS if c in df_accumulated.columns]
    df_accumulated[existing_cols].to_csv(LIVE_STREAM_PATH, index=False, encoding='utf-8-sig')


def main():
    start_time = time.time()
    print("=" * 65)
    print("🚀 JOB-FLOW AUTOMATOR — HIGH-PERFORMANCE LIVE PIPELINE")
    print("=" * 65, flush=True)

    # 1. Load historical exclusion sets (microsecond matching)
    hist_urls, hist_title_comp = load_all_historical_reviewed_keys()
    print(f"📚 Loaded {len(hist_urls)} historical URLs and {len(hist_title_comp)} (Title+Company) reviewed keys.")
    print("-" * 65, flush=True)

    all_jobs_master = []
    accumulated_live_df = pd.DataFrame()

    # 2. Fetch alternative VC boards first (super fast direct APIs ~2-3 sec)
    print("\n⚡ [STAGE 1/3] FETCHING FAST VC PORTFOLIOS (Ashby, Greenhouse, Arbeitnow)...", flush=True)
    df_custom = fetch_all_custom_boards()
    if not df_custom.empty:
        df_custom['category'] = 'VC Funds & Local Boards'
        df_custom['search_query'] = 'direct_api'
        all_jobs_master.append(df_custom)

        # Process and write first live batch
        df_custom['clean_url'] = df_custom['job_url'].apply(clean_job_url)
        df_custom['norm_title'] = df_custom['title'].apply(normalize_text_for_dedup)
        df_custom['norm_comp'] = df_custom['company'].apply(normalize_text_for_dedup)
        df_custom['dedup_key'] = df_custom['norm_title'] + "_" + df_custom['norm_comp']

        valid_custom = df_custom[
            (~df_custom['clean_url'].isin(hist_urls)) &
            (~df_custom['dedup_key'].isin(hist_title_comp)) &
            (df_custom['title'].apply(is_suitable_title)) &
            (df_custom.apply(lambda r: config.is_valid_location(r.get('location', ''), r.get('title', '')), axis=1)) &
            (~df_custom['description'].apply(requires_excluded_language))
        ].copy()

        if not valid_custom.empty:
            valid_custom['contact_email'] = valid_custom['description'].apply(extract_emails)
            valid_custom = apply_job_scoring(valid_custom)
            accumulated_live_df = valid_custom.sort_values(by='match_score', ascending=False)
            update_live_stream_file(accumulated_live_df)
            print(f"🟢 [LIVE STREAM READY] {len(accumulated_live_df)} instant jobs published to Live Feed!", flush=True)

    # 3. Scrape Main Job Boards with incremental stream updates
    print("\n🔍 [STAGE 2/3] SCRAPING MAIN BOARDS (LinkedIn, Indeed)...", flush=True)
    total_queries = sum(len(q) for q in config.SEARCH_CATEGORIES.values())
    current_query_num = 0

    for category_name, queries in config.SEARCH_CATEGORIES.items():
        print(f"\n📂 CATEGORY: {category_name.upper()}", flush=True)
        for query in queries:
            current_query_num += 1
            print(f"  [{current_query_num}/{total_queries}] 🔎 Searching: '{query}'...", end="", flush=True)
            try:
                jobs = scrape_jobs(
                    site_name=config.SEARCH_SITES,
                    search_term=query,
                    google_search_term=f"{query} jobs in {config.LOCATION}",
                    location=config.LOCATION,
                    results_wanted=config.RESULTS_PER_QUERY,
                    hours_old=config.HOURS_OLD,
                    country_indeed=config.COUNTRY_INDEED
                )
                if not jobs.empty:
                    jobs['category'] = category_name
                    jobs['search_query'] = query
                    all_jobs_master.append(jobs)
                    
                    counts = jobs['site'].str.lower().value_counts().to_dict()
                    breakdown = ", ".join([f"{site.capitalize()}: {count}" for site, count in counts.items()])
                    print(f"\n     ✓ Found: {len(jobs)} jobs ({breakdown})", flush=True)
                else:
                    print(f"\n     - Found: 0 jobs", flush=True)
            except Exception as e:
                print(f"\n     ⚠️ Error fetching '{query}': {e}", flush=True)
            
            time.sleep(random.uniform(1.5, 3.0))

    if not all_jobs_master:
        print("❌ No jobs found. Exiting.", flush=True)
        return

    # 3. Final Aggregation & Multithreaded Enrichment
    print("\n🧹 [STAGE 3/3] DEDUPLICATION, FAST ENRICHMENT & FINAL SCORING...", flush=True)
    df_raw = pd.concat(all_jobs_master, ignore_index=True)
    raw_total = len(df_raw)

    df_raw['clean_url'] = df_raw['job_url'].apply(clean_job_url)
    df_raw['norm_title'] = df_raw['title'].apply(normalize_text_for_dedup)
    df_raw['norm_company'] = df_raw['company'].apply(normalize_text_for_dedup)
    df_raw['dedup_key'] = df_raw['norm_title'] + "_" + df_raw['norm_company']

    # Cross-site and in-batch deduplication
    df_dedup = df_raw.drop_duplicates(subset=['clean_url'], keep='first')
    df_dedup = df_dedup.drop_duplicates(subset=['dedup_key'], keep='first')

    # Instant historical filter + Title filter + Location filter
    df_filtered = df_dedup[
        (~df_dedup['clean_url'].isin(hist_urls)) &
        (~df_dedup['dedup_key'].isin(hist_title_comp)) &
        (df_dedup['title'].apply(is_suitable_title)) &
        (df_dedup.apply(lambda r: config.is_valid_location(r.get('location', ''), r.get('title', '')), axis=1))
    ].copy()

    # Fast 5-thread LinkedIn enrichment
    df_filtered = enrich_descriptions_multithreaded(df_filtered)

    # Language regex filter, emails & scoring
    df_final = df_filtered[~df_filtered['description'].apply(requires_excluded_language)].copy()
    df_final['contact_email'] = df_final['description'].apply(extract_emails)
    df_final = apply_job_scoring(df_final)
    df_final = df_final.sort_values(by='match_score', ascending=False)

    print(f"\n📊 Summary:")
    print(f"   • Total raw postings fetched: {raw_total}")
    print(f"   • Previously reviewed / duplicates excluded: {raw_total - len(df_final)}")
    print(f"   • 🌟 BRAND NEW QUALIFIED JOBS: {len(df_final)}")

    # 5. Export Final Timestamped CSV & Live Stream
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    output_filename = f"jobs_clean_{timestamp}.csv"
    output_path = os.path.join(BASE_DIR, output_filename)

    existing_cols = [c for c in config.OUTPUT_COLUMNS if c in df_final.columns]
    df_final[existing_cols].to_csv(output_path, index=False, encoding='utf-8-sig')
    df_final[existing_cols].to_csv(LIVE_STREAM_PATH, index=False, encoding='utf-8-sig')

    elapsed = round(time.time() - start_time, 1)
    print(f"\n🎉 DONE in {elapsed}s! Saved as:\n   📁 {output_filename}", flush=True)


LOCK_FILE_PATH = os.path.join(BASE_DIR, ".scraper.lock")

if __name__ == "__main__":
    try:
        with open(LOCK_FILE_PATH, "w") as f:
            f.write(str(os.getpid()))
        main()
    finally:
        if os.path.exists(LOCK_FILE_PATH):
            try:
                os.remove(LOCK_FILE_PATH)
            except Exception:
                pass
